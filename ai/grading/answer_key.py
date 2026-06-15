"""
Parse a teacher-uploaded Excel answer key into a structured dict.

Expected Excel format (single sheet, row 1 = header):

    | question | type | answer |
    |----------|------|--------|
    | 1        | MCQ  | A      |
    | 2        | MCQ  | C      |
    | 1        | TF   | T      |   <- sequential T/F sub-question number
    | 2        | TF   | F      |
    | 1        | NUM  | 42     |

Types (case-insensitive): MCQ, TF, NUM
MCQ answers: A / B / C / D
TF answers:  T / True / Đ / Đúng  OR  F / False / S / Sai
NUM answers: any numeric string (comma or dot as decimal separator)
"""

import io
import openpyxl


def load_answer_key(path_or_bytes) -> dict:
    """Parse an Excel answer key file.

    Args:
        path_or_bytes: file path (str/Path) or raw bytes from an uploaded file.

    Returns:
        {
            "MCQ": {1: "A", 2: "C", ...},
            "TF":  {1: True, 2: False, ...},
            "NUM": {1: "42", ...},
        }

    Raises:
        ValueError: with a user-readable message describing the problem.
    """
    if isinstance(path_or_bytes, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(path_or_bytes), data_only=True)
    else:
        wb = openpyxl.load_workbook(str(path_or_bytes), data_only=True)

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Excel file is empty.")

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    required = {"question", "type", "answer"}
    missing = required - set(header)
    if missing:
        raise ValueError(
            f"Missing required column(s): {', '.join(sorted(missing))}. "
            f"Expected columns: question, type, answer."
        )

    q_idx = header.index("question")
    t_idx = header.index("type")
    a_idx = header.index("answer")

    result: dict = {"MCQ": {}, "TF": {}, "NUM": {}}
    seen: set = set()

    for row_num, row in enumerate(rows[1:], start=2):
        q_raw = row[q_idx]
        t_raw = row[t_idx]
        a_raw = row[a_idx]

        if q_raw is None and t_raw is None and a_raw is None:
            continue

        try:
            q = int(q_raw)
        except (TypeError, ValueError):
            raise ValueError(
                f"Row {row_num}: 'question' must be an integer, got {q_raw!r}."
            )
        if q <= 0:
            raise ValueError(
                f"Row {row_num}: 'question' must be positive, got {q}."
            )

        if t_raw is None:
            raise ValueError(f"Row {row_num}: 'type' is missing.")
        t = str(t_raw).strip().upper()
        if t not in {"MCQ", "TF", "NUM"}:
            raise ValueError(
                f"Row {row_num}: 'type' must be MCQ, TF, or NUM, got {t_raw!r}."
            )

        key = (q, t)
        if key in seen:
            raise ValueError(
                f"Row {row_num}: duplicate entry for question {q} type {t}."
            )
        seen.add(key)

        # Blank answer = teacher left this question out of the key (skip silently).
        if a_raw is None or str(a_raw).strip() == "":
            continue
        a_str = str(a_raw).strip()

        if t == "MCQ":
            a_norm = a_str.upper()
            if a_norm not in {"A", "B", "C", "D"}:
                raise ValueError(
                    f"Row {row_num}: MCQ answer must be A/B/C/D, got {a_raw!r}."
                )
            result["MCQ"][q] = a_norm

        elif t == "TF":
            a_lower = a_str.lower()
            if a_lower in {"t", "true", "đ", "đúng", "1"}:
                result["TF"][q] = True
            elif a_lower in {"f", "false", "s", "sai", "0"}:
                result["TF"][q] = False
            else:
                raise ValueError(
                    f"Row {row_num}: TF answer must be T/True or F/False, got {a_raw!r}."
                )

        else:  # NUM
            result["NUM"][q] = a_str

    if not any(result.values()):
        raise ValueError("Answer key is empty — no valid rows found.")

    return result
