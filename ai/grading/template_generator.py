"""
Generate a pre-filled Excel answer key template from an OMR result.

The template has three columns: question | type | answer.
question and type are filled from the detected layout; answer is left blank
for the teacher to fill in offline and upload via the frontend Phase-A flow.

Numbering exactly mirrors flatten_omr() in compare.py:
  MCQ  — global question numbers already stored in the answers dict keys
  TF   — sequential sub-question integers across all Câu groups (sorted by Câu number)
  NUM  — Câu number used directly as the question number
"""

import re
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment


def _cau_number(key: str) -> int | None:
    """Extract the integer from a 'Câu X' string."""
    m = re.search(r'\d+', str(key))
    return int(m.group()) if m else None


def generate_answer_key_template(omr_json: dict, output_path: "str | Path") -> Path:
    """Write a pre-filled answer key template Excel file from an OMR result dict.

    Args:
        omr_json:    Result dict from process_image() / read_sheet.py.
                     Must contain a "sections" key with at least one of
                     mcq_region, true_false_region, or numeric_region.
        output_path: Destination .xlsx path (created or overwritten).

    Returns:
        Resolved Path to the written file.

    Raises:
        ValueError: if no gradable sections are found in omr_json.
    """
    output_path = Path(output_path)
    sections = omr_json.get("sections", {})

    rows: list[tuple[int, str]] = []  # (question_number, type_string)

    # ── MCQ ──────────────────────────────────────────────────────────────────
    # mcq_region is a list of region dicts; answers keys are global question
    # number strings already remapped by _finalize_mcq_numbering.
    mcq_qs: list[int] = []
    for region in sections.get("mcq_region", []):
        for q_str in region.get("answers", {}):
            try:
                mcq_qs.append(int(q_str))
            except (ValueError, TypeError):
                continue
    for q in sorted(set(mcq_qs)):
        rows.append((q, "MCQ"))

    # ── TF ───────────────────────────────────────────────────────────────────
    # After _finalize_tf_numbering, true_false_region is a dict:
    #   {"Câu 41": {"a": ..., "b": ..., "c": ..., "d": ...}, ...}
    # Sub-questions are numbered sequentially across Câu groups sorted by
    # Câu number — exactly as flatten_omr() does.
    tf_section = sections.get("true_false_region", {})
    if isinstance(tf_section, dict) and tf_section:
        sub_num = 1
        letter_order = ["a", "b", "c", "d"]
        for cau_key in sorted(tf_section.keys(), key=lambda k: (_cau_number(k) or 0)):
            group = tf_section[cau_key]
            for letter in letter_order:
                if letter in group:
                    rows.append((sub_num, "TF"))
                    sub_num += 1

    # ── NUM ──────────────────────────────────────────────────────────────────
    # After _finalize_numeric_numbering, numeric_region is a dict:
    #   {"Câu 49": value_or_None, "Câu 50": value_or_None}
    # The Câu number IS the question number (as used by flatten_omr).
    num_section = sections.get("numeric_region", {})
    num_qs: list[int] = []
    if isinstance(num_section, dict):
        for cau_key in num_section:
            n = _cau_number(cau_key)
            if n is not None:
                num_qs.append(n)
    for n in sorted(set(num_qs)):
        rows.append((n, "NUM"))

    if not rows:
        raise ValueError(
            "No gradable sections (MCQ, TF, NUM) found in the OMR JSON. "
            "Run the pipeline on a valid exam sheet first."
        )

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Answer Key"

    # Header row
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for col, label in enumerate(["question", "type", "answer"], start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Instruction comment on the "answer" header so teachers know the format
    ws["C1"].comment = Comment(
        "Fill in the 'answer' column before uploading.\n"
        "MCQ : A / B / C / D\n"
        "TF  : T (true / Đúng) or F (false / Sai)\n"
        "NUM : numeric value, e.g. 3  or  8.5  or  06",
        "Azota Template Generator",
    )

    # Data rows — answer cell intentionally left blank (None → empty cell)
    type_fill = {
        "MCQ": PatternFill("solid", fgColor="E2EFDA"),
        "TF":  PatternFill("solid", fgColor="FFF2CC"),
        "NUM": PatternFill("solid", fgColor="FCE4D6"),
    }
    for q, t in rows:
        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=q).fill = type_fill[t]
        ws.cell(row=row_idx, column=2, value=t).fill = type_fill[t]
        ws.cell(row=row_idx, column=3, value=None)   # blank — teacher fills this

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 18

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
